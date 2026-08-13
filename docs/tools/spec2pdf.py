#!/usr/bin/env python3
"""DB 테이블 명세서 xlsx → A4 세로 PDF.

    python3 docs/tools/spec2pdf.py docs/DB_테이블_명세서_템플릿_운영스케줄.xlsx

출력은 같은 경로에 .pdf 확장자로 저장된다(두 번째 인자로 지정 가능).

[왜 Numbers/Excel 내보내기를 쓰지 않나]
Numbers는 xlsx의 페이지 설정(세로/A4)을 무시하고 시트 크기에 맞춘 커스텀 페이지로
내보낸다(예: 962x792 가로). openpyxl로 orientation=portrait를 넣어도 마찬가지다.
그래서 xlsx를 HTML로 옮긴 뒤 Chrome 헤드리스로 A4 세로 인쇄한다.

[전제] Google Chrome 설치. xlsx가 원본이므로 엑셀을 고친 뒤 이 스크립트만 다시 돌리면 된다.
"""
import html
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

CSS = """
@page { size: A4 portrait; margin: 14mm 12mm; }
* { box-sizing: border-box; }
body { font-family: -apple-system, "Apple SD Gothic Neo", "Malgun Gothic", sans-serif;
       font-size: 8pt; color: #1a1a1a; margin: 0; }
section { page-break-after: always; }
section:last-child { page-break-after: auto; }
h2 { font-size: 13pt; margin: 0 0 2mm; padding-bottom: 1.5mm;
     border-bottom: 1.2pt solid #2f5496; color: #1f3864; }
table { width: 100%; border-collapse: collapse; table-layout: fixed; }
td { border: 0.5pt solid #bfbfbf; padding: 1.1mm 1.4mm; vertical-align: middle;
     word-break: break-word; overflow-wrap: anywhere; line-height: 1.35; }
td.b { font-weight: 700; }
td.c { text-align: center; }
td.mono { font-family: "SF Mono", Menlo, monospace; font-size: 7.2pt; }
tr.spacer td { border: none; height: 2.2mm; padding: 0; }
"""

# 엑셀 채우기색 → 인쇄용 색
FILL_MAP = {"D9E2F3": "#dbe5f1", "EDEDED": "#efefef", "FFF2CC": "#fff2cc"}


def col_widths(ws, ncols):
    """엑셀 컬럼 폭을 비율로 환산. 미지정 컬럼은 기본 8.43.

    가로(엑셀)보다 좁은 A4 세로에 넣으면 'NULL'·'DATETIME2' 같은 짧은 값까지
    글자 단위로 쪼개진다. 좁은 컬럼에 하한을 주고 셀 여백만큼 상수를 더해 보정한다.
    """
    raw = []
    for i in range(1, ncols + 1):
        d = ws.column_dimensions.get(get_column_letter(i))
        w = d.width if d and d.width else 8.43
        raw.append(max(w, 6) + 4)
    total = sum(raw)
    pct = [w / total * 100 for w in raw]

    FLOOR = 6.2  # 헤더가 글자 단위로 쪼개지지 않는 최소 폭(%)
    for i, p in enumerate(pct):
        if p < FLOOR:
            widest = pct.index(max(pct))
            pct[widest] -= FLOOR - p
            pct[i] = FLOOR
    return pct


def build(ws):
    ncols = ws.max_column
    # 병합: 좌상단 셀에 colspan/rowspan, 나머지는 건너뛴다
    span, covered = {}, set()
    for rng in ws.merged_cells.ranges:
        span[(rng.min_row, rng.min_col)] = (rng.max_col - rng.min_col + 1,
                                            rng.max_row - rng.min_row + 1)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))

    out = ["<section>", f"<h2>{html.escape(ws.title)}</h2>", "<table>", "<colgroup>"]
    for w in col_widths(ws, ncols):
        out.append(f'<col style="width:{w:.2f}%">')
    out.append("</colgroup>")

    for r in range(1, ws.max_row + 1):
        cells = [ws.cell(r, c) for c in range(1, ncols + 1)]
        if not any(c.value is not None for c in cells):
            out.append(f'<tr class="spacer"><td colspan="{ncols}"></td></tr>')
            continue

        # 엑셀에서 옆 칸으로 넘쳐 보이던 단독 셀(제목·구획 헤더)은 HTML에선 행 전체를 쓰게 한다
        filled = [c for c in range(1, ncols + 1)
                  if ws.cell(r, c).value is not None and (r, c) not in covered]
        solo = len(filled) == 1 and filled[0] == 1 and (r, 1) not in span

        out.append("<tr>")
        for c in range(1, ncols + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(r, c)
            cs, rs = span.get((r, c), (1, 1))
            if solo:
                cs = ncols
            cls, style = [], []
            if cell.font and cell.font.b:
                cls.append("b")
            if cell.alignment and cell.alignment.horizontal == "center":
                cls.append("c")
            if cell.font and cell.font.name == "Menlo":
                cls.append("mono")
            rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.patternType else None
            if isinstance(rgb, str):
                bg = FILL_MAP.get(rgb[-6:].upper())
                if bg:
                    style.append(f"background:{bg}")
            if cell.font and cell.font.sz and cell.font.sz >= 14:
                style.append("font-size:12pt;border:none")
            elif cell.font and cell.font.sz == 11 and cell.font.b:
                style.append("font-size:9.5pt;border:none;padding-top:3mm")
            v = "" if cell.value is None else html.escape(str(cell.value))
            attrs = f' colspan="{cs}"' if cs > 1 else ""
            attrs += f' rowspan="{rs}"' if rs > 1 else ""
            attrs += f' class="{" ".join(cls)}"' if cls else ""
            attrs += f' style="{";".join(style)}"' if style else ""
            out.append(f"<td{attrs}>{v}</td>")
            if solo:
                break
        out.append("</tr>")
    out.append("</table></section>")
    return "\n".join(out)


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1]).resolve()
    dst = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else src.with_suffix(".pdf")
    if not src.exists():
        sys.exit(f"엑셀 파일을 찾을 수 없습니다: {src}")
    if not Path(CHROME).exists():
        sys.exit(f"Google Chrome이 필요합니다: {CHROME}")

    wb = openpyxl.load_workbook(src)
    doc = ["<!doctype html><html><head><meta charset='utf-8'>",
           f"<title>{html.escape(src.stem)}</title>",
           f"<style>{CSS}</style></head><body>"]
    doc += [build(ws) for ws in wb.worksheets]
    doc.append("</body></html>")

    with tempfile.NamedTemporaryFile("w", suffix=".html", encoding="utf-8", delete=False) as f:
        f.write("\n".join(doc))
        tmp = f.name

    subprocess.run([CHROME, "--headless", "--disable-gpu", "--no-pdf-header-footer",
                    f"--print-to-pdf={dst}", f"file://{tmp}"],
                   check=True, capture_output=True)
    Path(tmp).unlink()
    print(f"생성 완료: {dst}  (시트 {len(wb.worksheets)}개)")


if __name__ == "__main__":
    main()
